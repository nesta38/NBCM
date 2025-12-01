"""
NBCM V2.5 - Service Rapports
Génération des rapports PDF et Excel professionnels
Version améliorée avec :
- Filtres Excel sur les en-têtes
- Anomalies en premier dans les PDF
- En-têtes répétés sur chaque page PDF
- Style harmonisé entre Rapport et Archivage
"""
import io
from datetime import datetime
from collections import defaultdict
from flask import current_app

from app.services.compliance_service import get_jobs_map, normalize_hostname


# =============================================================================
# STYLES COMMUNS
# =============================================================================

def get_excel_styles():
    """Retourne les styles Excel réutilisables."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    return {
        'header_font': Font(bold=True, color="FFFFFF", size=11),
        'header_fill_ok': PatternFill("solid", fgColor="198754"),
        'header_fill_ko': PatternFill("solid", fgColor="dc3545"),
        'header_fill_out': PatternFill("solid", fgColor="FF8C00"),
        'header_fill_blue': PatternFill("solid", fgColor="4472C4"),
        'header_fill_summary': PatternFill("solid", fgColor="1F4E78"),
        'row_fill_ok': PatternFill("solid", fgColor="E2EFDA"),
        'row_fill_ko': PatternFill("solid", fgColor="FCE4D6"),
        'row_fill_out': PatternFill("solid", fgColor="FFF3CD"),
        'thin_border': Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        ),
        'center_align': Alignment(horizontal="center", vertical="center"),
        'left_align': Alignment(horizontal="left", vertical="center"),
        'title_font': Font(size=16, bold=True, color="1F4E78"),
        'subtitle_font': Font(size=11, italic=True, color="666666"),
        'stat_font': Font(bold=True, size=11),
        'stat_value_font': Font(bold=True, size=12, color="FFFFFF"),
    }


def setup_excel_sheet_filters(ws, last_col, last_row):
    """Configure les filtres automatiques sur une feuille Excel."""
    from openpyxl.utils import get_column_letter
    if last_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_row}"


def adjust_column_widths(ws, widths_dict):
    """Ajuste les largeurs de colonnes."""
    for col_letter, width in widths_dict.items():
        ws.column_dimensions[col_letter].width = width


# =============================================================================
# RAPPORT EXCEL - PAGE RAPPORT
# =============================================================================

def generate_excel_report(conformite):
    """
    Génère un rapport Excel détaillé avec filtres sur les en-têtes.
    """
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        
        wb = openpyxl.Workbook()
        jobs_map = get_jobs_map()
        styles = get_excel_styles()
        
        # === SHEET 1: Résumé ===
        ws = wb.active
        ws.title = "Résumé"
        
        # Titre
        ws['A1'] = "NetBackup Compliance Report"
        ws['A1'].font = styles['title_font']
        ws.merge_cells('A1:D1')
        
        ws['A2'] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        ws['A2'].font = styles['subtitle_font']
        ws.merge_cells('A2:D2')
        
        # Statistiques avec couleurs
        stats = [
            ("Taux de Conformité", f"{conformite['taux_conformite']}%", 
             styles['header_fill_ok'] if conformite['taux_conformite'] >= 95 else styles['header_fill_ko']),
            ("Serveurs Attendus (Backup=Oui)", conformite['total_attendus'], styles['header_fill_blue']),
            ("Serveurs Conformes", conformite['conformes'], styles['header_fill_ok']),
            ("Serveurs Non Conformes", conformite['non_conformes'], styles['header_fill_ko']),
            ("Serveurs Hors CMDB", conformite['non_references'], styles['header_fill_out']),
            ("Jobs Analysés (24h)", conformite['total_jobs'], styles['header_fill_blue']),
        ]
        
        for i, (label, value, fill) in enumerate(stats, 4):
            ws[f'A{i}'] = label
            ws[f'A{i}'].font = styles['stat_font']
            ws[f'A{i}'].border = styles['thin_border']
            ws[f'B{i}'] = value
            ws[f'B{i}'].fill = fill
            ws[f'B{i}'].font = styles['stat_value_font']
            ws[f'B{i}'].alignment = styles['center_align']
            ws[f'B{i}'].border = styles['thin_border']
        
        adjust_column_widths(ws, {'A': 35, 'B': 15})
        
        # === SHEET 2: Non Conformes (Anomalies en premier) ===
        if conformite['liste_non_conformes']:
            ws_ko = wb.create_sheet("⚠ Non Conformes")
            
            headers_ko = ['Hostname', 'Statut', 'Diagnostic', 'Action Requise']
            for col, val in enumerate(headers_ko, 1):
                cell = ws_ko.cell(row=1, column=col)
                cell.value = val
                cell.font = styles['header_font']
                cell.fill = styles['header_fill_ko']
                cell.alignment = styles['center_align']
                cell.border = styles['thin_border']
            
            row_num = 2
            for host in conformite['liste_non_conformes']:
                ws_ko.cell(row=row_num, column=1).value = host
                ws_ko.cell(row=row_num, column=2).value = "NON CONFORME"
                ws_ko.cell(row=row_num, column=3).value = "Aucun backup valide dans les 24 dernières heures"
                ws_ko.cell(row=row_num, column=4).value = "Vérifier la configuration NetBackup"
                
                for c in range(1, 5):
                    ws_ko.cell(row=row_num, column=c).fill = styles['row_fill_ko']
                    ws_ko.cell(row=row_num, column=c).border = styles['thin_border']
                row_num += 1
            
            # Filtres automatiques
            setup_excel_sheet_filters(ws_ko, 4, row_num - 1)
            adjust_column_widths(ws_ko, {'A': 30, 'B': 18, 'C': 45, 'D': 35})
        
        # === SHEET 3: Hors CMDB ===
        if conformite['liste_non_references']:
            ws_out = wb.create_sheet("⚡ Hors CMDB")
            
            headers_out = ['Hostname', 'Date Backup', 'Policy', 'Schedule', 'Statut Job', 'Taille (GB)']
            for col, val in enumerate(headers_out, 1):
                cell = ws_out.cell(row=1, column=col)
                cell.value = val
                cell.font = styles['header_font']
                cell.fill = styles['header_fill_out']
                cell.alignment = styles['center_align']
                cell.border = styles['thin_border']
            
            row_num = 2
            for host in conformite['liste_non_references']:
                norm_name = normalize_hostname(host)
                jobs = jobs_map.get(norm_name, [])
                
                if jobs:
                    for job in jobs:
                        ws_out.cell(row=row_num, column=1).value = host
                        ws_out.cell(row=row_num, column=2).value = job.backup_time.strftime('%d/%m/%Y %H:%M') if job.backup_time else ''
                        ws_out.cell(row=row_num, column=3).value = job.policy_name or ''
                        ws_out.cell(row=row_num, column=4).value = job.schedule_name or ''
                        ws_out.cell(row=row_num, column=5).value = job.status or ''
                        ws_out.cell(row=row_num, column=6).value = round(job.taille_gb, 2) if job.taille_gb else 0
                        
                        for c in range(1, 7):
                            ws_out.cell(row=row_num, column=c).fill = styles['row_fill_out']
                            ws_out.cell(row=row_num, column=c).border = styles['thin_border']
                        row_num += 1
                else:
                    ws_out.cell(row=row_num, column=1).value = host
                    ws_out.cell(row=row_num, column=2).value = "Non trouvé"
                    for c in range(1, 7):
                        ws_out.cell(row=row_num, column=c).fill = styles['row_fill_out']
                        ws_out.cell(row=row_num, column=c).border = styles['thin_border']
                    row_num += 1
            
            setup_excel_sheet_filters(ws_out, 6, row_num - 1)
            adjust_column_widths(ws_out, {'A': 30, 'B': 18, 'C': 40, 'D': 20, 'E': 12, 'F': 12})
        
        # === SHEET 4: Détail Complet (Conformes) ===
        ws_detail = wb.create_sheet("✓ Conformes - Détail")
        
        headers = ['Hostname', 'Statut Global', 'Date Backup', 'Job ID', 
                   'Policy', 'Schedule', 'Taille (GB)', 'Durée (min)', 'Statut Job']
        
        for col, val in enumerate(headers, 1):
            cell = ws_detail.cell(row=1, column=col)
            cell.value = val
            cell.font = styles['header_font']
            cell.fill = styles['header_fill_ok']
            cell.alignment = styles['center_align']
            cell.border = styles['thin_border']
        
        row_num = 2
        for host_cmdb in conformite['liste_conformes']:
            norm_name = normalize_hostname(host_cmdb)
            jobs = jobs_map.get(norm_name, [])
            
            if jobs:
                for job in jobs:
                    ws_detail.cell(row=row_num, column=1).value = host_cmdb
                    ws_detail.cell(row=row_num, column=2).value = "CONFORME"
                    ws_detail.cell(row=row_num, column=3).value = job.backup_time.strftime('%d/%m/%Y %H:%M') if job.backup_time else ''
                    ws_detail.cell(row=row_num, column=4).value = job.job_id or ''
                    ws_detail.cell(row=row_num, column=5).value = job.policy_name or ''
                    ws_detail.cell(row=row_num, column=6).value = job.schedule_name or ''
                    ws_detail.cell(row=row_num, column=7).value = round(job.taille_gb, 2) if job.taille_gb else 0
                    ws_detail.cell(row=row_num, column=8).value = job.duree_minutes or 0
                    ws_detail.cell(row=row_num, column=9).value = job.status or ''
                    
                    for c in range(1, 10):
                        ws_detail.cell(row=row_num, column=c).fill = styles['row_fill_ok']
                        ws_detail.cell(row=row_num, column=c).border = styles['thin_border']
                    row_num += 1
            else:
                ws_detail.cell(row=row_num, column=1).value = host_cmdb
                ws_detail.cell(row=row_num, column=2).value = "CONFORME"
                ws_detail.cell(row=row_num, column=3).value = "Job archivé"
                for c in [4, 5, 6, 7, 8]:
                    ws_detail.cell(row=row_num, column=c).value = "-"
                ws_detail.cell(row=row_num, column=9).value = "OK"
                for c in range(1, 10):
                    ws_detail.cell(row=row_num, column=c).fill = styles['row_fill_ok']
                    ws_detail.cell(row=row_num, column=c).border = styles['thin_border']
                row_num += 1
        
        # Filtres automatiques
        setup_excel_sheet_filters(ws_detail, 9, row_num - 1)
        adjust_column_widths(ws_detail, {
            'A': 28, 'B': 15, 'C': 18, 'D': 15, 
            'E': 40, 'F': 20, 'G': 12, 'H': 12, 'I': 12
        })
        
        # Figer la première ligne sur toutes les feuilles
        for sheet in wb.worksheets:
            if sheet.title != "Résumé":
                sheet.freeze_panes = 'A2'
        
        # Sauvegarder
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
        
    except Exception as e:
        current_app.logger.error(f"Erreur génération Excel: {e}", exc_info=True)
        return None


# =============================================================================
# RAPPORT PDF - PAGE RAPPORT
# =============================================================================

def generate_pdf_report(conformite):
    """
    Génère un rapport PDF professionnel.
    - Anomalies en premier après le résumé
    - En-têtes répétés sur chaque page
    - Codes couleur appropriés
    """
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.units import cm
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=landscape(A4),
            leftMargin=1*cm, 
            rightMargin=1*cm,
            topMargin=1.5*cm,
            bottomMargin=1.5*cm
        )
        
        elements = []
        styles = getSampleStyleSheet()
        jobs_map = get_jobs_map()
        
        # Couleurs
        COLOR_OK = colors.HexColor('#198754')
        COLOR_KO = colors.HexColor('#dc3545')
        COLOR_OUT = colors.HexColor('#FF8C00')
        COLOR_BLUE = colors.HexColor('#4472C4')
        COLOR_LIGHT_OK = colors.HexColor('#E2EFDA')
        COLOR_LIGHT_KO = colors.HexColor('#FCE4D6')
        COLOR_LIGHT_OUT = colors.HexColor('#FFF3CD')
        
        # Styles personnalisés
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            spaceAfter=20,
            textColor=colors.HexColor('#1F4E78')
        )
        
        section_style = ParagraphStyle(
            'SectionTitle',
            parent=styles['Heading2'],
            fontSize=14,
            spaceBefore=15,
            spaceAfter=10,
            textColor=colors.HexColor('#1F4E78')
        )
        
        # === PAGE 1: Titre et Résumé ===
        elements.append(Paragraph(
            f"NetBackup Compliance Report", 
            title_style
        ))
        elements.append(Paragraph(
            f"<font size='10' color='#666666'>Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')} - Période analysée: 24 dernières heures</font>",
            styles['Normal']
        ))
        elements.append(Spacer(1, 1*cm))
        
        # Taux de conformité mis en avant - CORRIGÉ: espacement et styles séparés
        rate_color = '#198754' if conformite['taux_conformite'] >= 95 else '#dc3545'
        rate_style = ParagraphStyle('RateValue', alignment=1, fontSize=48, leading=50)
        elements.append(Paragraph(
            f"<font color='{rate_color}'><b>{conformite['taux_conformite']}%</b></font>",
            rate_style
        ))
        elements.append(Spacer(1, 0.3*cm))
        label_style = ParagraphStyle('RateLabel', alignment=1, fontSize=14)
        elements.append(Paragraph(
            f"Taux de Conformité Global",
            label_style
        ))
        elements.append(Spacer(1, 1*cm))
        
        # Tableau récapitulatif
        summary_data = [
            ['Métrique', 'Valeur', 'Statut'],
            ['Serveurs Attendus (Backup=Oui)', str(conformite['total_attendus']), 'Périmètre CMDB'],
            ['Serveurs Conformes', str(conformite['conformes']), '✓ OK'],
            ['Serveurs Non Conformes', str(conformite['non_conformes']), '✗ Action requise' if conformite['non_conformes'] > 0 else '✓ OK'],
            ['Serveurs Hors CMDB', str(conformite['non_references']), '⚡ À vérifier' if conformite['non_references'] > 0 else '✓ OK'],
            ['Jobs Analysés (24h)', str(conformite['total_jobs']), 'Traités'],
        ]
        
        summary_table = Table(summary_data, colWidths=[10*cm, 4*cm, 6*cm])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), COLOR_BLUE),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            # Couleurs spécifiques pour les valeurs
            ('BACKGROUND', (1, 2), (1, 2), COLOR_OK),  # Conformes
            ('TEXTCOLOR', (1, 2), (1, 2), colors.white),
            ('BACKGROUND', (1, 3), (1, 3), COLOR_KO if conformite['non_conformes'] > 0 else COLOR_OK),
            ('TEXTCOLOR', (1, 3), (1, 3), colors.white),
            ('BACKGROUND', (1, 4), (1, 4), COLOR_OUT if conformite['non_references'] > 0 else COLOR_OK),
            ('TEXTCOLOR', (1, 4), (1, 4), colors.white),
        ]))
        elements.append(summary_table)
        
        # === SECTION ANOMALIES (Non Conformes) - EN PREMIER ===
        if conformite['liste_non_conformes']:
            elements.append(PageBreak())
            elements.append(Paragraph(
                f"<font color='#dc3545'>⚠ ANOMALIES - Serveurs Non Conformes ({len(conformite['liste_non_conformes'])})</font>", 
                section_style
            ))
            elements.append(Paragraph(
                "<font size='10' color='#dc3545'><b>Action immédiate requise - Ces serveurs n'ont pas de backup valide dans les 24 dernières heures</b></font>",
                styles['Normal']
            ))
            elements.append(Spacer(1, 0.3*cm))
            
            data_ko = [['#', 'Hostname', 'Diagnostic', 'Action Requise']]
            for idx, h in enumerate(conformite['liste_non_conformes'], 1):
                data_ko.append([
                    str(idx), 
                    h, 
                    'Aucun backup valide (24h)', 
                    'Vérifier configuration NetBackup'
                ])
            
            table_ko = Table(data_ko, colWidths=[1*cm, 8*cm, 8*cm, 8*cm], repeatRows=1)
            table_ko.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), COLOR_KO),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [COLOR_LIGHT_KO, colors.white]),
            ]))
            elements.append(table_ko)
        
        # === SECTION HORS CMDB ===
        if conformite['liste_non_references']:
            elements.append(PageBreak())
            elements.append(Paragraph(
                f"<font color='#FF8C00'>⚡ Serveurs Hors CMDB ({len(conformite['liste_non_references'])})</font>", 
                section_style
            ))
            elements.append(Paragraph(
                "<font size='10'>Ces serveurs effectuent des backups mais ne sont pas référencés dans la CMDB</font>",
                styles['Normal']
            ))
            elements.append(Spacer(1, 0.3*cm))
            
            data_out = [['#', 'Hostname', 'Date Backup', 'Policy', 'Statut']]
            for idx, host in enumerate(conformite['liste_non_references'], 1):
                norm_name = normalize_hostname(host)
                jobs = jobs_map.get(norm_name, [])
                if jobs:
                    job = jobs[0]
                    data_out.append([
                        str(idx),
                        host[:35],
                        job.backup_time.strftime('%d/%m/%Y %H:%M') if job.backup_time else '-',
                        (job.policy_name or '')[:35],
                        job.status or '-'
                    ])
                else:
                    data_out.append([str(idx), host[:35], '-', '-', '-'])
            
            table_out = Table(data_out, colWidths=[1*cm, 8*cm, 5*cm, 9*cm, 2*cm], repeatRows=1)
            table_out.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), COLOR_OUT),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                ('ALIGN', (-1, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [COLOR_LIGHT_OUT, colors.white]),
            ]))
            elements.append(table_out)
        
        # === SECTION CONFORMES - TOUS LES SERVEURS ===
        if conformite['liste_conformes']:
            elements.append(PageBreak())
            elements.append(Paragraph(
                f"<font color='#198754'>✓ Serveurs Conformes ({len(conformite['liste_conformes'])})</font>", 
                section_style
            ))
            elements.append(Spacer(1, 0.3*cm))
            
            data_ok = [['#', 'Hostname', 'Date Backup', 'Policy', 'Statut']]
            count = 0
            for hostname in conformite['liste_conformes']:
                norm_name = normalize_hostname(hostname)
                jobs = jobs_map.get(norm_name, [])
                count += 1
                if jobs:
                    job = jobs[0]
                    data_ok.append([
                        str(count),
                        hostname[:35],
                        job.backup_time.strftime('%d/%m/%Y %H:%M') if job.backup_time else '-',
                        (job.policy_name or '')[:35],
                        job.status or 'OK'
                    ])
                else:
                    data_ok.append([str(count), hostname[:35], '-', 'Archivé', 'OK'])
            
            table_ok = Table(data_ok, colWidths=[1*cm, 8*cm, 5*cm, 9*cm, 2*cm], repeatRows=1)
            table_ok.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), COLOR_OK),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                ('ALIGN', (-1, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [COLOR_LIGHT_OK, colors.white]),
            ]))
            elements.append(table_ok)
        
        doc.build(elements)
        buffer.seek(0)
        return buffer
        
    except Exception as e:
        current_app.logger.error(f"Erreur génération PDF: {e}", exc_info=True)
        return None


# =============================================================================
# RAPPORT EXCEL - ARCHIVAGE (Harmonisé avec page Rapport)
# =============================================================================

def generate_excel_report_archive(conformite, archive):
    """
    Génère un Excel pour une archive - même format que le rapport standard.
    """
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        
        wb = openpyxl.Workbook()
        styles = get_excel_styles()
        
        date_debut = archive.date_debut_periode.strftime('%d/%m/%Y %Hh')
        date_fin = archive.date_fin_periode.strftime('%d/%m/%Y %Hh')
        
        # === SHEET 1: Résumé ===
        ws = wb.active
        ws.title = "Résumé"
        
        ws['A1'] = f"Archive NetBackup - {date_debut} → {date_fin}"
        ws['A1'].font = styles['title_font']
        ws.merge_cells('A1:D1')
        
        ws['A2'] = f"Archivé le {archive.date_archivage.strftime('%d/%m/%Y à %H:%M')}"
        ws['A2'].font = styles['subtitle_font']
        ws.merge_cells('A2:D2')
        
        stats = [
            ("Période", f"{date_debut} → {date_fin}", styles['header_fill_blue']),
            ("Taux de Conformité", f"{conformite['taux_conformite']}%", 
             styles['header_fill_ok'] if conformite['taux_conformite'] >= 95 else styles['header_fill_ko']),
            ("Serveurs Attendus", conformite.get('total_attendus', conformite.get('total_backup_enabled', 0)), styles['header_fill_blue']),
            ("Serveurs Conformes", len(conformite['liste_conformes']), styles['header_fill_ok']),
            ("Serveurs Non Conformes", len(conformite['liste_non_conformes']), styles['header_fill_ko']),
            ("Serveurs Hors CMDB", len(conformite['liste_non_references']), styles['header_fill_out']),
            ("Total Jobs", conformite['total_jobs'], styles['header_fill_blue']),
        ]
        
        for i, (label, value, fill) in enumerate(stats, 4):
            ws[f'A{i}'] = label
            ws[f'A{i}'].font = styles['stat_font']
            ws[f'A{i}'].border = styles['thin_border']
            ws[f'B{i}'] = value
            ws[f'B{i}'].fill = fill
            ws[f'B{i}'].font = styles['stat_value_font']
            ws[f'B{i}'].alignment = styles['center_align']
            ws[f'B{i}'].border = styles['thin_border']
        
        adjust_column_widths(ws, {'A': 35, 'B': 25})
        
        # === SHEET 2: Non Conformes ===
        if conformite['liste_non_conformes']:
            ws_ko = wb.create_sheet("⚠ Non Conformes")
            
            headers_ko = ['#', 'Hostname', 'Diagnostic']
            for col, val in enumerate(headers_ko, 1):
                cell = ws_ko.cell(row=1, column=col)
                cell.value = val
                cell.font = styles['header_font']
                cell.fill = styles['header_fill_ko']
                cell.alignment = styles['center_align']
                cell.border = styles['thin_border']
            
            for idx, h in enumerate(conformite['liste_non_conformes'], 1):
                row_num = idx + 1
                ws_ko.cell(row=row_num, column=1).value = idx
                ws_ko.cell(row=row_num, column=2).value = h
                ws_ko.cell(row=row_num, column=3).value = "Aucun backup valide pendant la période"
                for c in range(1, 4):
                    ws_ko.cell(row=row_num, column=c).fill = styles['row_fill_ko']
                    ws_ko.cell(row=row_num, column=c).border = styles['thin_border']
            
            setup_excel_sheet_filters(ws_ko, 3, len(conformite['liste_non_conformes']) + 1)
            adjust_column_widths(ws_ko, {'A': 8, 'B': 35, 'C': 45})
            ws_ko.freeze_panes = 'A2'
        
        # === SHEET 3: Hors CMDB ===
        if conformite['liste_non_references']:
            ws_out = wb.create_sheet("⚡ Hors CMDB")
            
            headers_out = ['#', 'Hostname', 'Remarque']
            for col, val in enumerate(headers_out, 1):
                cell = ws_out.cell(row=1, column=col)
                cell.value = val
                cell.font = styles['header_font']
                cell.fill = styles['header_fill_out']
                cell.alignment = styles['center_align']
                cell.border = styles['thin_border']
            
            for idx, h in enumerate(conformite['liste_non_references'], 1):
                row_num = idx + 1
                ws_out.cell(row=row_num, column=1).value = idx
                ws_out.cell(row=row_num, column=2).value = h
                ws_out.cell(row=row_num, column=3).value = "Backup effectué mais serveur non référencé CMDB"
                for c in range(1, 4):
                    ws_out.cell(row=row_num, column=c).fill = styles['row_fill_out']
                    ws_out.cell(row=row_num, column=c).border = styles['thin_border']
            
            setup_excel_sheet_filters(ws_out, 3, len(conformite['liste_non_references']) + 1)
            adjust_column_widths(ws_out, {'A': 8, 'B': 35, 'C': 50})
            ws_out.freeze_panes = 'A2'
        
        # === SHEET 4: Conformes ===
        if conformite['liste_conformes']:
            ws_ok = wb.create_sheet("✓ Conformes")
            
            headers_ok = ['#', 'Hostname', 'Statut']
            for col, val in enumerate(headers_ok, 1):
                cell = ws_ok.cell(row=1, column=col)
                cell.value = val
                cell.font = styles['header_font']
                cell.fill = styles['header_fill_ok']
                cell.alignment = styles['center_align']
                cell.border = styles['thin_border']
            
            for idx, h in enumerate(conformite['liste_conformes'], 1):
                row_num = idx + 1
                ws_ok.cell(row=row_num, column=1).value = idx
                ws_ok.cell(row=row_num, column=2).value = h
                ws_ok.cell(row=row_num, column=3).value = "CONFORME"
                for c in range(1, 4):
                    ws_ok.cell(row=row_num, column=c).fill = styles['row_fill_ok']
                    ws_ok.cell(row=row_num, column=c).border = styles['thin_border']
            
            setup_excel_sheet_filters(ws_ok, 3, len(conformite['liste_conformes']) + 1)
            adjust_column_widths(ws_ok, {'A': 8, 'B': 35, 'C': 15})
            ws_ok.freeze_panes = 'A2'
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
        
    except Exception as e:
        current_app.logger.error(f"Erreur Excel archive: {e}", exc_info=True)
        return None


# =============================================================================
# RAPPORT PDF - ARCHIVAGE (Harmonisé avec page Rapport)
# =============================================================================

def generate_pdf_report_archive(conformite, archive):
    """
    Génère un PDF pour une archive - même style que le rapport standard.
    - Anomalies en premier après le résumé
    - En-têtes répétés sur chaque page
    """
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.units import cm
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=landscape(A4), 
            leftMargin=1*cm, 
            rightMargin=1*cm,
            topMargin=1.5*cm,
            bottomMargin=1.5*cm
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Couleurs
        COLOR_OK = colors.HexColor('#198754')
        COLOR_KO = colors.HexColor('#dc3545')
        COLOR_OUT = colors.HexColor('#FF8C00')
        COLOR_BLUE = colors.HexColor('#4472C4')
        COLOR_LIGHT_OK = colors.HexColor('#E2EFDA')
        COLOR_LIGHT_KO = colors.HexColor('#FCE4D6')
        COLOR_LIGHT_OUT = colors.HexColor('#FFF3CD')
        
        date_debut = conformite['date_debut_periode'].strftime('%d/%m/%Y %Hh')
        date_fin = conformite['date_fin_periode'].strftime('%d/%m/%Y %Hh')
        
        # Styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=15,
            textColor=colors.HexColor('#1F4E78')
        )
        
        section_style = ParagraphStyle(
            'SectionTitle',
            parent=styles['Heading2'],
            fontSize=14,
            spaceBefore=15,
            spaceAfter=10,
            textColor=colors.HexColor('#1F4E78')
        )
        
        # === PAGE 1: Titre et Résumé ===
        elements.append(Paragraph(
            f"Archive NetBackup", 
            title_style
        ))
        elements.append(Paragraph(
            f"<font size='12'><b>Période: {date_debut} → {date_fin}</b></font>",
            styles['Normal']
        ))
        elements.append(Paragraph(
            f"<font size='10' color='#666666'>Archivé le {archive.date_archivage.strftime('%d/%m/%Y à %H:%M')}</font>",
            styles['Normal']
        ))
        elements.append(Spacer(1, 1*cm))
        
        # Taux de conformité - CORRIGÉ: espacement et styles séparés
        rate_color = '#198754' if conformite['taux_conformite'] >= 95 else '#dc3545'
        rate_style = ParagraphStyle('RateValue', alignment=1, fontSize=48, leading=50)
        elements.append(Paragraph(
            f"<font color='{rate_color}'><b>{conformite['taux_conformite']}%</b></font>",
            rate_style
        ))
        elements.append(Spacer(1, 0.3*cm))
        label_style = ParagraphStyle('RateLabel', alignment=1, fontSize=14)
        elements.append(Paragraph(
            f"Taux de Conformité",
            label_style
        ))
        elements.append(Spacer(1, 1*cm))
        
        # Tableau récapitulatif
        summary_data = [
            ['Métrique', 'Valeur'],
            ['Période couverte', f"{date_debut} → {date_fin}"],
            ['Serveurs Attendus', str(conformite.get('total_attendus', conformite.get('total_backup_enabled', 0)))],
            ['Serveurs Conformes', str(len(conformite['liste_conformes']))],
            ['Serveurs Non Conformes', str(len(conformite['liste_non_conformes']))],
            ['Serveurs Hors CMDB', str(len(conformite['liste_non_references']))],
            ['Total Jobs', str(conformite['total_jobs'])],
        ]
        
        summary_table = Table(summary_data, colWidths=[10*cm, 6*cm])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), COLOR_BLUE),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            # Couleurs spécifiques
            ('BACKGROUND', (1, 3), (1, 3), COLOR_OK),
            ('TEXTCOLOR', (1, 3), (1, 3), colors.white),
            ('BACKGROUND', (1, 4), (1, 4), COLOR_KO if len(conformite['liste_non_conformes']) > 0 else COLOR_OK),
            ('TEXTCOLOR', (1, 4), (1, 4), colors.white),
            ('BACKGROUND', (1, 5), (1, 5), COLOR_OUT if len(conformite['liste_non_references']) > 0 else COLOR_OK),
            ('TEXTCOLOR', (1, 5), (1, 5), colors.white),
        ]))
        elements.append(summary_table)
        
        # === SECTION ANOMALIES (Non Conformes) - EN PREMIER ===
        if conformite['liste_non_conformes']:
            elements.append(PageBreak())
            elements.append(Paragraph(
                f"<font color='#dc3545'>⚠ ANOMALIES - Serveurs Non Conformes ({len(conformite['liste_non_conformes'])})</font>", 
                section_style
            ))
            elements.append(Paragraph(
                "<font size='10' color='#dc3545'><b>Ces serveurs n'avaient pas de backup valide pendant la période archivée</b></font>",
                styles['Normal']
            ))
            elements.append(Spacer(1, 0.3*cm))
            
            data_ko = [['#', 'Hostname', 'Diagnostic']]
            for idx, h in enumerate(conformite['liste_non_conformes'], 1):
                data_ko.append([str(idx), h, 'Aucun backup valide pendant la période'])
            
            table_ko = Table(data_ko, colWidths=[1*cm, 10*cm, 14*cm], repeatRows=1)
            table_ko.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), COLOR_KO),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [COLOR_LIGHT_KO, colors.white]),
            ]))
            elements.append(table_ko)
        
        # === SECTION HORS CMDB ===
        if conformite['liste_non_references']:
            elements.append(PageBreak())
            elements.append(Paragraph(
                f"<font color='#FF8C00'>⚡ Serveurs Hors CMDB ({len(conformite['liste_non_references'])})</font>", 
                section_style
            ))
            elements.append(Paragraph(
                "<font size='10'>Serveurs ayant effectué des backups mais non référencés dans la CMDB</font>",
                styles['Normal']
            ))
            elements.append(Spacer(1, 0.3*cm))
            
            data_out = [['#', 'Hostname', 'Remarque']]
            for idx, h in enumerate(conformite['liste_non_references'], 1):
                data_out.append([str(idx), h, 'Non référencé dans CMDB'])
            
            table_out = Table(data_out, colWidths=[1*cm, 12*cm, 12*cm], repeatRows=1)
            table_out.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), COLOR_OUT),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [COLOR_LIGHT_OUT, colors.white]),
            ]))
            elements.append(table_out)
        
        # === SECTION CONFORMES ===
        if conformite['liste_conformes']:
            elements.append(PageBreak())
            elements.append(Paragraph(
                f"<font color='#198754'>✓ Serveurs Conformes ({len(conformite['liste_conformes'])})</font>", 
                section_style
            ))
            elements.append(Spacer(1, 0.3*cm))
            
            data_ok = [['#', 'Hostname', 'Statut']]
            for idx, h in enumerate(conformite['liste_conformes'], 1):
                data_ok.append([str(idx), h, 'CONFORME'])
            
            table_ok = Table(data_ok, colWidths=[1*cm, 14*cm, 10*cm], repeatRows=1)
            table_ok.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), COLOR_OK),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [COLOR_LIGHT_OK, colors.white]),
            ]))
            elements.append(table_ok)
        
        doc.build(elements)
        buffer.seek(0)
        return buffer
        
    except Exception as e:
        current_app.logger.error(f"Erreur PDF archive: {e}", exc_info=True)
        return None
